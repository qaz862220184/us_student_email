import asyncio
import pyppeteer
import openpyxl as xl
import os
from lxml import etree
from pyppeteer import launch
import re

email = 'qifcz5e6@temporary-mail.net'

def screen_size():
    # 使用tkinter获取屏幕大小
    import tkinter
    tk = tkinter.Tk()
    width = tk.winfo_screenwidth()
    height = tk.winfo_screenheight()
    tk.quit()
    return width, height



async def get_edu(email_detail):

    browser = await launch({
        'executablePath': pyppeteer.launcher.executablePath(),
        #'headless': False,
        'dumpio': True,
        'autoClose': True,
        'args': [
            '--no-sandbox',
            "--start-maximized",
            '--disable-infobars',
            '--disable-extensions',
            '--hide-scrollbars',
            '--disable-bundled-ppapi-flash',
            '--mute-audio',
            '--no-sandbox',
            '--disable-setuid-sandbox',
            '--disable-gpu',
            '--enable-automation',
            #'--proxy-server={}'.format(proxy_ip)
            #'--proxy-server=192.168.1.250:7890',
            '--proxy-server=192.168.1.239:24002'
        ]
    })
    page = await browser.newPage()
    width, height = screen_size()
    # print(width,height)
    await page.setViewport({  # 最大化窗口
        "width": width,
        "height": 900
    })
    await page.evaluateOnNewDocument('() =>{ Object.defineProperties(navigator,'
                                     '{ webdriver:{ get: () => false } }) }')



    # 获取email地址
    url1 = 'https://www.temporary-mail.net/mailbox/'+ email_detail
    #await page.goto('https://www.temporary-mail.net/change')
    await page.goto(url1)
    await page.waitFor(10000)

    email_href = ''
    contents = await page.content()
    tree = etree.HTML(contents)
    tr_list = tree.xpath('//*[@id="message-list"]/tr')
    print(tr_list)
    for tr in tr_list:
        email_title = tr.xpath('./td[2]//text()')[0]
        #email_sender = tr.xpath('./td[1]/a/text()')[0]
        #print(email_title,email_sender)
        #sender = 'NO_Reply'
        if email_title == 'WELCOME TO MERCED COLLEGE':
            print('就是这封邮件')
            email_href = tr.xpath('./td[2]/a/@href')[0]
            break

    if email_href =='':
        return 0
    else:
        email_url = 'https://www.temporary-mail.net' + email_href
        print(email_url)
        await page.goto(email_url)
        contents1 = await page.content()
        tree1 = etree.HTML(contents1)
        text_part = ''
        detail = tree1.xpath('//*[@id="main"]/div/div/div[2]/div[1]/div/div[3]/text()')
        for data in detail:
            text_part += data

        stu_id = re.search('\d{7}',text_part).group()
        print(stu_id)

        # edu_email = re.findall(r'([\w-]+(\.[\w-]+)*@[\w-]+(\.[\w-]+)+)',text_part)
        # print(edu_email)

        edu_email = re.search(r'([\w-]+(\.[\w-]+)*(@campus\.mccd\.edu))', text_part).group()
        print(edu_email)
        #print(text_part)

        return stu_id,edu_email





        ## 接下来就是从detail中取到学生id和邮箱地址

def write_excel_file(folder_path,data):
    result_path = os.path.join(folder_path, "edu_email.xlsx")
    print(result_path)

    print('***** 开始写入excel文件 ' + result_path + ' ***** \n')
    #if os.path.exists(result_path):
        #print('***** excel已存在，在表后添加数据 ' + result_path + ' ***** \n')
    workbook = xl.Workbook()
    workbook.save(result_path)
    workbook = xl.load_workbook(result_path)
    sheet = workbook.active
    for key in data:
        print(key)
        sheet.append(key)
    workbook.save(result_path)

    print('***** 生成Excel文件 ' + result_path + ' ***** \n')

def red_excel_file(folder_path):
    result_path = os.path.join(folder_path, "edu_zhuce.xlsx")
    wb2 = xl.load_workbook(result_path)
    sheet = wb2.active

    for row in sheet.rows:
        user_detail1 = []
        for cell in row:
            user_detail1.append(cell.value)
        user_data.append(user_detail1)


if __name__ == '__main__':
    user_data = []
    red_excel_file('./')
    user_data[0].append("stu_id")
    user_data[0].append('edu_email')
    user_data[0].append('edu_pwd')
    for i in range(1,len(user_data)):
        a = user_data[i][-1].split('@')[0]
        Birthday = user_data[i][5]
        print(Birthday)
        m,d,y = Birthday.split('/')
        m=m.zfill(2)
        d=d.zfill(2)
        y=y[-2:]
        edu_pwd=m+d+y

        #print(a)
    #print(user_data[1][-1])

        stu_id,edu_email = asyncio.get_event_loop().run_until_complete(get_edu(a))
        #print(stu_id,edu_email)
        user_data[i].append(stu_id)
        user_data[i].append(edu_email)
        user_data[i].append(edu_pwd)
        #print(user_data[i])

    for data in user_data:
        print(data)
    write_excel_file('./',user_data)

